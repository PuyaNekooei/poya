from django.contrib import admin
from django.db.models import Sum, Count
from django.http import HttpResponse
from import_export import resources, fields
from import_export.admin import ImportExportModelAdmin
from import_export.formats import base_formats
from .models import Order, OrderItem, Table
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import jdatetime


# Resource classes for import/export
class OrderResource(resources.ModelResource):
    order_number = fields.Field(column_name='شماره سفارش', attribute='order_number')
    customer_name = fields.Field(column_name='نام مشتری', attribute='customer_name')
    customer_phone = fields.Field(column_name='شماره تماس مشتری', attribute='customer_phone')
    total_amount = fields.Field(column_name='مبلغ کل', attribute='total_amount')
    status = fields.Field(column_name='وضعیت', attribute='status')
    order_type = fields.Field(column_name='نوع سفارش', attribute='order_type')
    notes = fields.Field(column_name='یادداشت', attribute='notes')
    created_by__username = fields.Field(column_name='ایجادکننده', attribute='created_by__username')
    created_at = fields.Field(column_name='تاریخ ایجاد', attribute='created_at')
    updated_at = fields.Field(column_name='تاریخ بروزرسانی', attribute='updated_at')
    
    class Meta:
        model = Order
        fields = ('id', 'order_number', 'customer_name', 'customer_phone', 
                  'total_amount', 'status', 'order_type', 'notes', 
                  'created_by__username', 'created_at', 'updated_at')
        export_order = ('id', 'order_number', 'customer_name', 'customer_phone', 
                        'total_amount', 'status', 'order_type', 'created_at')


class OrderItemResource(resources.ModelResource):
    item_name = fields.Field(column_name='نام محصول')
    order__order_number = fields.Field(column_name='شماره سفارش', attribute='order__order_number')
    menu_item__name = fields.Field(column_name='آیتم منو', attribute='menu_item__name')
    quantity = fields.Field(column_name='تعداد', attribute='quantity')
    unit_price = fields.Field(column_name='قیمت واحد', attribute='unit_price')
    total_price = fields.Field(column_name='قیمت کل', attribute='total_price')
    notes = fields.Field(column_name='یادداشت', attribute='notes')
    created_at = fields.Field(column_name='تاریخ ایجاد', attribute='created_at')
    
    class Meta:
        model = OrderItem
        fields = ('id', 'order__order_number', 'item_name', 'menu_item__name', 
                  'quantity', 'unit_price', 'total_price', 'notes', 'created_at')
        export_order = ('id', 'order__order_number', 'item_name', 
                        'quantity', 'unit_price', 'total_price', 'created_at')
    
    def dehydrate_item_name(self, obj):
        if obj.menu_item:
            return obj.menu_item.name
        return obj.product_info.get('product_name', 'Unknown')


# Inline admin classes
class OrderItemInline(admin.TabularInline):
    model = OrderItem
    extra = 1
    readonly_fields = ['total_price']


# Admin classes with import/export functionality
@admin.register(Order)
class OrderAdmin(ImportExportModelAdmin):
    resource_class = OrderResource
    list_display = ['order_number', 'customer_name', 'table', 'total_amount', 'status', 'order_type', 'created_at']
    list_filter = ['status', 'order_type', 'table', 'created_at']
    search_fields = ['order_number', 'customer_name', 'customer_phone']
    readonly_fields = ['order_number', 'total_amount', 'created_at', 'updated_at']
    inlines = [OrderItemInline]
    ordering = ['-created_at']
    actions = ['export_daily_sales_report']
    
    def export_daily_sales_report(self, request, queryset):
        """Export daily sales report showing quantity and revenue by date and menu item"""
        # Get all order items from selected orders
        order_items = OrderItem.objects.filter(
            order__in=queryset
        ).select_related('order', 'menu_item', 'menu_item__category')
        
        # Group by date and menu item
        from collections import defaultdict
        daily_sales = defaultdict(lambda: defaultdict(lambda: {
            'quantity': 0, 
            'revenue': Decimal('0.00'), 
            'menu_name': '', 
            'category': '', 
            'product_code': '',
            'price_without_tax': Decimal('0.00'),
            'price_with_tax': Decimal('0.00'),
            'vat_rate': 0
        }))
        
        for item in order_items:
            date = item.order.created_at.date()
            if item.menu_item:
                menu_id = item.menu_item.id
                menu_name = item.menu_item.name
                category = item.menu_item.category.name if item.menu_item.category else ''
                product_code = item.menu_item.product_code or ''
                price_without_tax = item.menu_item.price_without_tax or Decimal('0.00')
                price_with_tax = item.menu_item.price_with_tax or Decimal('0.00')
                vat_rate = item.menu_item.vat_rate
            else:
                menu_id = f"unknown_{item.id}"
                menu_name = item.product_info.get('product_name', 'محصول ناشناخته')
                category = ''
                product_code = ''
                price_without_tax = Decimal('0.00')
                price_with_tax = item.unit_price or Decimal('0.00')
                vat_rate = 0
            
            daily_sales[date][menu_id]['menu_name'] = menu_name
            daily_sales[date][menu_id]['category'] = category
            daily_sales[date][menu_id]['product_code'] = product_code
            daily_sales[date][menu_id]['price_without_tax'] = price_without_tax
            daily_sales[date][menu_id]['price_with_tax'] = price_with_tax
            daily_sales[date][menu_id]['vat_rate'] = vat_rate
            daily_sales[date][menu_id]['quantity'] += item.quantity
            daily_sales[date][menu_id]['revenue'] += item.total_price
        
        # Create Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'گزارش فروش روزانه'
        
        # Header style
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers
        headers = ['تاریخ شمسی', 'نام غذا', 'کد محصول', 'دسته‌بندی', 
                   'قیمت بدون مالیات', 'درصد مالیات', 'قیمت با مالیات', 'تعداد فروش', 'مجموع درآمد']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Data rows
        row = 2
        for date in sorted(daily_sales.keys(), reverse=True):
            # Convert to Jalali date
            jalali_date = jdatetime.date.fromgregorian(date=date)
            jalali_str = jalali_date.strftime('%Y/%m/%d')
            
            for menu_id, data in sorted(daily_sales[date].items(), key=lambda x: x[1]['menu_name']):
                ws.cell(row=row, column=1, value=jalali_str)
                ws.cell(row=row, column=2, value=data['menu_name'])
                ws.cell(row=row, column=3, value=data['product_code'])
                ws.cell(row=row, column=4, value=data['category'])
                ws.cell(row=row, column=5, value=float(data['price_without_tax']))
                ws.cell(row=row, column=6, value=f"{data['vat_rate']}%")
                ws.cell(row=row, column=7, value=float(data['price_with_tax']))
                ws.cell(row=row, column=8, value=data['quantity'])
                ws.cell(row=row, column=9, value=float(data['revenue']))
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15  # تاریخ شمسی
        ws.column_dimensions['B'].width = 30  # نام غذا
        ws.column_dimensions['C'].width = 15  # کد محصول
        ws.column_dimensions['D'].width = 20  # دسته‌بندی
        ws.column_dimensions['E'].width = 18  # قیمت بدون مالیات
        ws.column_dimensions['F'].width = 12  # درصد مالیات
        ws.column_dimensions['G'].width = 18  # قیمت با مالیات
        ws.column_dimensions['H'].width = 15  # تعداد فروش
        ws.column_dimensions['I'].width = 20  # مجموع درآمد
        
        # Save to response
        from io import BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Use Jalali date for filename
        jalali_now = jdatetime.datetime.now()
        filename = f'daily_sales_report_{jalali_now.strftime("%Y%m%d")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    export_daily_sales_report.short_description = 'گزارش فروش روزانه (Excel)'


@admin.register(OrderItem)
class OrderItemAdmin(ImportExportModelAdmin):
    resource_class = OrderItemResource
    list_display = ['order', 'get_item_name', 'quantity', 'unit_price', 'total_price', 'has_product_info']
    list_filter = ['order__status', 'order__order_type', 'created_at']
    search_fields = ['order__order_number', 'menu_item__name', 'product_info']
    readonly_fields = ['total_price']
    ordering = ['-created_at']
    
    def get_item_name(self, obj):
        if obj.menu_item:
            return obj.menu_item.name
        else:
            return obj.product_info.get('product_name', 'Unknown')
    get_item_name.short_description = 'Item Name'
    
    def has_product_info(self, obj):
        return bool(obj.product_info)
    has_product_info.boolean = True
    has_product_info.short_description = 'Has Product Info'


@admin.register(Table)
class TableAdmin(admin.ModelAdmin):
    list_display = ['number', 'name', 'capacity', 'status', 'is_active']
    list_filter = ['status', 'is_active']
    search_fields = ['number', 'name']
    list_editable = ['status', 'is_active']
    ordering = ['number']
